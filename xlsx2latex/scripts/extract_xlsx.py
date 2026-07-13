#!/usr/bin/env python3
"""
Extract structure from an .xlsx table into compact JSON for LaTeX conversion.

Reads cell values, merges, bold/italic/alignment as *exceptions to a default*
(not a full per-cell dump), flags narrow empty columns as spacers, formats
numbers the way Excel displays them (not raw floats), and detects horizontal
border rows as hints for where booktabs rules (\\toprule/\\midrule/\\bottomrule)
belong.

Usage:
    python extract_xlsx.py table.xlsx
    python extract_xlsx.py table.xlsx --sheet "Results" --range A1:F12
    python extract_xlsx.py table.xlsx --out structure.json
"""
import argparse
import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

SPACER_WIDTH_THRESHOLD = 3.5  # Excel column-width units; narrower than this + empty => spacer
BOLD_ROW_MAJORITY = 0.6       # fraction of non-empty cells in a row that must be bold to call the row bold


def format_number(value, number_format):
    """Render a numeric cell the way Excel displays it (decimals, %, thousands sep),
    not the raw float. Returns (display_string, is_percent, decimals)."""
    if number_format is None or number_format == "General":
        # Excel's default float display: trim trailing zeros
        if isinstance(value, float) and value == int(value):
            return str(int(value)), False, 0
        return repr(value).rstrip("0").rstrip(".") if isinstance(value, float) else str(value), False, None

    fmt = number_format
    is_percent = "%" in fmt
    v = value * 100 if is_percent else value

    # count decimal places from the format pattern, e.g. "0.00" -> 2, "0.000%" -> 3
    m = re.search(r"0\.([0#]+)", fmt)
    decimals = len(m.group(1)) if m else 0

    has_thousands = "," in fmt.split(".")[0] if "." in fmt else "," in fmt

    try:
        if has_thousands:
            display = f"{v:,.{decimals}f}"
        else:
            display = f"{v:.{decimals}f}"
    except (TypeError, ValueError):
        return str(value), is_percent, decimals

    if is_percent:
        display += "%"
    return display, is_percent, decimals


def cell_alignment(cell):
    h = cell.alignment.horizontal if cell.alignment else None
    if h in (None, "general"):
        return None
    return h


def extract(path, sheet_name=None, cell_range=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_col, min_row = 1, 1
        max_col, max_row = ws.max_column, ws.max_row

    n_rows = max_row - min_row + 1
    n_cols = max_col - min_col + 1

    # --- column widths & spacer detection ---
    columns = []
    col_has_content = [False] * n_cols
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                col_has_content[c - min_col] = True

    for c in range(min_col, max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        width = dim.width if dim and dim.width else 8.43
        idx = c - min_col
        is_spacer = (width < SPACER_WIDTH_THRESHOLD) and not col_has_content[idx]
        columns.append({"index": idx, "width": round(width, 2), "is_spacer": is_spacer})

    # --- merges (within range only) ---
    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row < min_row or mr.max_row > max_row or mr.min_col < min_col or mr.max_col > max_col:
            continue
        merges.append({
            "row0": mr.min_row - min_row, "col0": mr.min_col - min_col,
            "row1": mr.max_row - min_row, "col1": mr.max_col - min_col,
            "rowspan": mr.max_row - mr.min_row + 1,
            "colspan": mr.max_col - mr.min_col + 1,
        })

    # --- rows: bold-row detection (majority rule) + per-cell exceptions ---
    bold_rows = []
    border_top_rows = []
    border_bottom_rows = []
    double_border_bottom_rows = []
    rows_out = []

    for r in range(min_row, max_row + 1):
        ridx = r - min_row
        row_cells = []
        n_nonempty = 0
        n_bold = 0
        row_has_top_border = False
        row_has_bottom_border = False
        row_has_double_bottom = False

        for c in range(min_col, max_col + 1):
            cidx = c - min_col
            if columns[cidx]["is_spacer"]:
                continue
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val in (None, ""):
                continue
            n_nonempty += 1
            is_bold = bool(cell.font and cell.font.bold)
            if is_bold:
                n_bold += 1

            entry = {"col": cidx}
            if isinstance(val, (int, float)):
                display, is_pct, decimals = format_number(val, cell.number_format)
                entry["value"] = val
                entry["display"] = display
                entry["numeric"] = True
                if is_pct:
                    entry["percent"] = True
                if decimals is not None:
                    entry["decimals"] = decimals
            else:
                text = str(val)
                entry["value"] = text
                entry["numeric"] = False

            if cell.font and cell.font.italic:
                entry["italic"] = True
            align = cell_alignment(cell)
            if align:
                entry["align"] = align

            row_cells.append(entry)

            # border check (top/bottom of this cell)
            b = cell.border
            if b:
                if b.top and b.top.style:
                    row_has_top_border = True
                if b.bottom and b.bottom.style:
                    row_has_bottom_border = True
                    if b.bottom.style == "double":
                        row_has_double_bottom = True

        if n_nonempty and n_bold / n_nonempty >= BOLD_ROW_MAJORITY:
            bold_rows.append(ridx)
            # strip redundant per-cell bold flags implied by the row default
        else:
            # mark isolated bold cells only when the row itself isn't bold
            for c in range(min_col, max_col + 1):
                cidx = c - min_col
                if columns[cidx]["is_spacer"]:
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value not in (None, "") and cell.font and cell.font.bold:
                    for entry in row_cells:
                        if entry["col"] == cidx:
                            entry["bold"] = True

        if row_has_top_border:
            border_top_rows.append(ridx)
        if row_has_bottom_border:
            border_bottom_rows.append(ridx)
        if row_has_double_bottom:
            double_border_bottom_rows.append(ridx)

        if row_cells:
            rows_out.append({"index": ridx, "cells": row_cells})

    result = {
        "source": str(path),
        "sheet": ws.title,
        "n_rows": n_rows,
        "n_cols": n_cols,
        "columns": columns,
        "merges": merges,
        "rows": rows_out,
        "style_notes": {
            "bold_rows": bold_rows,
            "spacer_cols": [c["index"] for c in columns if c["is_spacer"]],
            "border_top_rows": border_top_rows,
            "border_bottom_rows": border_bottom_rows,
            "double_border_bottom_rows": double_border_bottom_rows,
        },
    }
    return result


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="Path to the .xlsx file")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    ap.add_argument("--range", dest="cell_range", default=None, help="Cell range e.g. A1:F12 (default: used range)")
    ap.add_argument("--out", default=None, help="Write JSON here instead of stdout")
    args = ap.parse_args()

    try:
        result = extract(args.xlsx, args.sheet, args.cell_range)
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)

    out_json = json.dumps(result, indent=2)
    if args.out:
        with open(args.out, "w") as f:
            f.write(out_json)
        print(f"Wrote {args.out} ({result['n_rows']} rows x {result['n_cols']} cols, "
              f"{len(result['style_notes']['spacer_cols'])} spacer col(s), "
              f"{len(result['merges'])} merge(s))", file=sys.stderr)
    else:
        print(out_json)


if __name__ == "__main__":
    main()
